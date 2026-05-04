#!/usr/bin/env python3
"""
Migra bases Excel de compañías a PostgreSQL en tablas separadas.

Ontario -> ontario_companies
Quebec  -> quebec_companies

Por defecto reinicia ambas tablas de importación para corregir cargas mezcladas.
Para no truncar, usar RESET_COMPANY_IMPORT_TABLES=false.
"""

import json
import os
import re
import sys
import unicodedata
from datetime import datetime
from pathlib import Path

import openpyxl
import psycopg2
from psycopg2.extras import execute_batch
from dotenv import load_dotenv

load_dotenv()

DB_URL = os.getenv("DATABASE_URL", "postgresql://casaos:casaos@127.0.0.1:5434/casaos")
RESET_TABLES = os.getenv("RESET_COMPANY_IMPORT_TABLES", "true").lower() not in {"0", "false", "no"}

DATASETS = [
    {
        "name": "Ontario",
        "table": "ontario_companies",
        "file": Path(r"C:\Users\ripre\Downloads\BASE DE DATOS PROVINCIA DE ONTARIO sheets.xlsx"),
    },
    {
        "name": "Quebec",
        "table": "quebec_companies",
        "file": Path(r"C:\Users\ripre\Downloads\Base De Datos Provincia De Quebec sheets.xlsx"),
    },
]

TARGET_FIELDS = [
    "nombre",
    "telefono",
    "tipo",
    "correo",
    "direccion",
    "provincia",
    "region",
    "ciudad",
    "pueblo",
    "work",
    "descripcion",
    "dominio_de_pagina",
    "lista_de_llamadas",
]

HEADER_ALIASES = {
    "nombre": "nombre",
    "telefono": "telefono",
    "tipo": "tipo",
    "correo": "correo",
    "direccion": "direccion",
    "provincia": "provincia",
    "region": "region",
    "ciudad": "ciudad",
    "pueblo": "pueblo",
    "work": "work",
    "descripcion": "descripcion",
    "descripcion del trabajo": "descripcion",
    "dominio de pagina": "dominio_de_pagina",
    "lista de llamadas": "lista_de_llamadas",
}


def clean_value(value):
    if value is None:
        return None
    cleaned = re.sub(r"\s+", " ", str(value)).strip()
    return cleaned or None


def normalize_name(name):
    cleaned = clean_value(name)
    return cleaned.lower() if cleaned else None


def normalize_header(value):
    cleaned = clean_value(value)
    if not cleaned:
        return ""
    cleaned = unicodedata.normalize("NFKD", cleaned)
    cleaned = "".join(ch for ch in cleaned if not unicodedata.combining(ch))
    return cleaned.lower()


def assert_table_name(table):
    allowed = {dataset["table"] for dataset in DATASETS}
    if table not in allowed:
        raise ValueError(f"Tabla no permitida: {table}")


def ensure_table(cursor, table):
    assert_table_name(table)
    cursor.execute('CREATE EXTENSION IF NOT EXISTS "uuid-ossp"')
    cursor.execute(
        f"""
        CREATE TABLE IF NOT EXISTS {table} (
            id UUID PRIMARY KEY DEFAULT uuid_generate_v4(),
            nombre TEXT NOT NULL,
            telefono TEXT,
            tipo TEXT,
            correo TEXT,
            direccion TEXT,
            provincia TEXT,
            region TEXT,
            ciudad TEXT,
            pueblo TEXT,
            work TEXT,
            descripcion TEXT,
            dominio_de_pagina TEXT,
            lista_de_llamadas TEXT,
            is_duplicate BOOLEAN DEFAULT FALSE,
            status VARCHAR(50) DEFAULT 'pending',
            created_at TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP
        )
        """
    )
    for column in [
        "nombre",
        "telefono",
        "tipo",
        "correo",
        "provincia",
        "region",
        "ciudad",
        "pueblo",
        "work",
        "dominio_de_pagina",
    ]:
        cursor.execute(f"ALTER TABLE {table} ALTER COLUMN {column} TYPE TEXT")
    cursor.execute(
        f"""
        CREATE UNIQUE INDEX IF NOT EXISTS idx_{table}_nombre_unique
        ON {table} (LOWER(TRIM(nombre)))
        WHERE is_duplicate = FALSE
        """
    )
    cursor.execute(
        f"""
        CREATE INDEX IF NOT EXISTS idx_{table}_created_at
        ON {table}(created_at DESC)
        """
    )


def read_excel_records(excel_file):
    wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
    ws = wb["Hoja 1"] if "Hoja 1" in wb.sheetnames else wb[wb.sheetnames[0]]

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    header_map = {}
    for index, header in enumerate(header_row):
        target = HEADER_ALIASES.get(normalize_header(header))
        if target:
            header_map[index] = target

    if "nombre" not in set(header_map.values()):
        raise ValueError(f"No se encontró columna de nombre en {excel_file}")

    records = []
    errors = []
    for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        record = {field: None for field in TARGET_FIELDS}
        for index, target in header_map.items():
            if index < len(row):
                record[target] = clean_value(row[index])

        if not record["nombre"]:
            if any(clean_value(value) for value in row):
                errors.append({"file": str(excel_file), "row": row_number, "error": "Nombre vacío"})
            continue

        records.append({"row": row_number, "data": record})

    return records, errors


def migrate_dataset(cursor, dataset):
    table = dataset["table"]
    excel_file = dataset["file"]
    ensure_table(cursor, table)

    records, errors = read_excel_records(excel_file)

    if RESET_TABLES:
        cursor.execute(f"TRUNCATE TABLE {table}")
        existing_names = set()
    else:
        cursor.execute(
            f"""
            SELECT LOWER(REGEXP_REPLACE(TRIM(nombre), '\\s+', ' ', 'g'))
            FROM {table}
            WHERE is_duplicate = FALSE
            """
        )
        existing_names = {row[0] for row in cursor.fetchall()}

    records_to_insert = []
    duplicates = []
    for item in records:
        record = item["data"]
        name_norm = normalize_name(record["nombre"])
        if name_norm in existing_names:
            duplicates.append({
                "row": item["row"],
                "nombre": record["nombre"],
                "reason": "Ya existe en la tabla o en el lote",
            })
            continue
        records_to_insert.append(tuple(record[field] for field in TARGET_FIELDS))
        existing_names.add(name_norm)

    if records_to_insert:
        execute_batch(
            cursor,
            f"""
            INSERT INTO {table}
            (nombre, telefono, tipo, correo, direccion, provincia, region, ciudad,
             pueblo, work, descripcion, dominio_de_pagina, lista_de_llamadas)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            ON CONFLICT DO NOTHING
            """,
            records_to_insert,
            page_size=1000,
        )

    cursor.execute(f"SELECT COUNT(*)::int FROM {table} WHERE is_duplicate = FALSE")
    active_total = cursor.fetchone()[0]

    return {
        "dataset": dataset["name"],
        "table": table,
        "file": str(excel_file),
        "rows": len(records),
        "inserted_attempted": len(records_to_insert),
        "duplicates": len(duplicates),
        "errors": len(errors),
        "active_total_after": active_total,
        "duplicates_sample": duplicates[:10],
        "errors_sample": errors[:10],
    }


def migrate_companies():
    print("=" * 80)
    print("MIGRACIÓN: Excel -> PostgreSQL (tablas separadas)")
    print("=" * 80)
    print(f"Reiniciar tablas antes de importar: {RESET_TABLES}")

    for dataset in DATASETS:
        if not dataset["file"].exists():
            print(f"Archivo no encontrado: {dataset['file']}")
            sys.exit(1)
        print(f"- {dataset['name']}: {dataset['file']} -> {dataset['table']}")

    print("\n[1/3] Conectando a PostgreSQL...")
    try:
        conn = psycopg2.connect(DB_URL)
        cursor = conn.cursor()
        print("  OK conexión")
    except Exception as exc:
        print(f"Error de conexión: {exc}")
        sys.exit(1)

    print("\n[2/3] Importando datasets...")
    results = []
    try:
        for dataset in DATASETS:
            result = migrate_dataset(cursor, dataset)
            results.append(result)
            print(
                f"  OK {result['dataset']}: {result['inserted_attempted']} insertadas/enviadas, "
                f"{result['duplicates']} duplicados, {result['errors']} errores, "
                f"total tabla {result['active_total_after']}"
            )
        conn.commit()
    except Exception as exc:
        conn.rollback()
        conn.close()
        print(f"Error en migración: {exc}")
        sys.exit(1)

    print("\n[3/3] Guardando reporte...")
    report = {
        "timestamp": datetime.now().isoformat(),
        "reset_tables": RESET_TABLES,
        "datasets": results,
        "total_inserted_attempted": sum(item["inserted_attempted"] for item in results),
        "total_duplicates": sum(item["duplicates"] for item in results),
        "total_errors": sum(item["errors"] for item in results),
    }
    report_file = Path("migration_report.json")
    report_file.write_text(json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8")
    conn.close()

    print("\n" + "=" * 80)
    print("REPORTE FINAL")
    print("=" * 80)
    for item in results:
        print(f"{item['dataset']} ({item['table']}): {item['active_total_after']} activos")
    print(f"Duplicados totales: {report['total_duplicates']}")
    print(f"Errores totales:    {report['total_errors']}")
    print(f"Reporte:            {report_file.resolve()}")


if __name__ == "__main__":
    migrate_companies()
