#!/usr/bin/env python3
import openpyxl
import json

excel_file = r"C:\Users\ripre\Downloads\BASE DE DATOS PROVINCIA DE ONTARIO sheets.xlsx"
wb = openpyxl.load_workbook(excel_file)

result = {}
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    headers = []
    for cell in ws[1]:
        headers.append(cell.value)
    
    result[sheet_name] = {
        "headers": headers,
        "max_row": ws.max_row,
        "max_col": ws.max_column,
        "sample_data": []
    }
    
    # Get first few rows
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=3, values_only=True)):
        result[sheet_name]["sample_data"].append(row)

print(json.dumps(result, indent=2, default=str))
